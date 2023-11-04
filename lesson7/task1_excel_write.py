from openpyxl import Workbook
from openpyxl.styles import Alignment

wordbook = Workbook()

for sheet_number in range(1, 11):
    wordbook.create_sheet(f"multiplication{sheet_number}")
    worksheet = wordbook[f"multiplication{sheet_number}"]

    for row in range(1, 101):
        worksheet.append((sheet_number, "*", row, "=", sheet_number * row))
        for column in range(1, worksheet.max_column+1):
            cell = worksheet.cell(row, column)
            cell.alignment = Alignment(horizontal="center", vertical="center")




wordbook.save("multiplication.xlsx")

