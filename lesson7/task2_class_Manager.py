from openpyxl import load_workbook
from openpyxl.cell import MergedCell, Cell
from typing import Union

TABLE_NAME = "price13.11.2022.xlsx"


class Manager:
    def __init__(self) -> None:
        self.workbook = load_workbook(TABLE_NAME)
        self.worksheet = self.workbook["zavod"]

    def _get_rows(self) -> list[Union[Cell, MergedCell]]:
        rows = []
        for row_index in range(2, self.worksheet.max_row):
            row = self.worksheet[f"A{row_index}":f"E{row_index}"][0]

            if MergedCell in map(type, row) or row[0].value is None:
                continue

            rows.append(row)

        return rows

    def print_price(self) -> None:
        for row in self._get_rows():
            print(row[1].value.strip(), "price is", row[3].value)

    def reserve_item(self, item_id: int, count_of_item_to_reserve: int = 0) -> bool:
        rows = self._get_rows()

        item_row = tuple(filter(lambda row: item_id == row[0].value, rows))[0]
        reserved = 0 if item_row[4].value is None else item_row[4].value
        available = int(item_row[2].value) - reserved
        print(f"already reserved: {reserved}, available products count =", available, "with price",
              item_row[3].value, "per item!")

        if available - count_of_item_to_reserve < 0:
            print("[WARN] Can't reserve that item, count of items to reserve is too big! Available items =", available)
            return False

        self.worksheet[item_row[4].coordinate] = reserved + count_of_item_to_reserve
        print(f"Items successfully reserved! {item_id}: {item_row[1].value}, {count_of_item_to_reserve} items!")

        return True

    def delete_order_item(self, item_id: int, amount: int) -> bool:
        rows = self._get_rows()

        item_row = tuple(filter(lambda row: item_id == row[0].value, rows))[0]
        reserved = 0 if item_row[4].value is None else item_row[4].value

        print("trying to delete item with id:", item_row[0].value, "and name:", item_row[1].value,
              "and amount:", amount)

        if reserved - amount < 0:
            raise ValueError(f"delete_order_item: item id:{item_id}. Value is too big, reserved: {reserved}")

        self.worksheet[item_row[4].coordinate] = None if reserved - amount == 0 else reserved - amount

        return True

    def projected_revenue(self) -> Union[float, int]:
        rows = self._get_rows()

        summary_revenue = 0
        for item_row in rows:
            reserved = 0 if item_row[4].value is None else item_row[4].value
            price = item_row[3].value

            summary_revenue += reserved * price

        summary_revenue = round(summary_revenue, 5)
        print("Total revenue is", summary_revenue)
        return summary_revenue


if __name__ == "__main__":
    manager = Manager()
    #
    # print("-----TEST-----")
    # manager.print_price()
    #
    # print("Is reservation success:", manager.reserve_item(1, 50))
    # print("Delete reservation 30:", manager.delete_order_item(1, 30))
    #
    # try:
    #     print("Delete reservation overload", manager.delete_order_item(1, 50))
    # except:
    #     print("Overload test OK!")
    #
    # print("Delete reservation remaining items (20)", manager.delete_order_item(1, 20))
    #
    # manager.projected_revenue()
    #
    # print("-----TEST_END-----")
    #
    # manager.workbook.save(TABLE_NAME)

    command = "A"
    while command != "":
        command = input("Enter command (help for help xD or Enter for exit) > ").lower()

        if "help" in command:
            print("printPrice - printing all items prices")
            print("reserveItem <item_id: int> <item_amount: int> - reserving item")
            print("deleteOrderItem <item_id: int> <order_amount: int> - deletes order!")
            print("projectedRevenue - printing total revenue")
        elif "printprice" in command:
            manager.print_price()
        elif "reserveitem" in command:
            command, item_id, amount = command.split()
            manager.reserve_item(int(item_id), int(amount))
        elif "deleteorderitem" in command:
            command, item_id, amount = command.split()
            manager.delete_order_item(int(item_id), int(amount))
        elif "projectedrevenue" in command:
            manager.projected_revenue()
        elif command == "":
            print("Good bye!")
        else:
            print("Unknown command! Enter help!")

        manager.workbook.save(TABLE_NAME)

