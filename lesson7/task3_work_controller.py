from os import path
from openpyxl import Workbook, load_workbook
from datetime import datetime
from time import sleep
from psutil import process_iter

if not path.exists("Manager'sWorkingDay.xlsx"):
    print(f"File with name Manager'sWorkingDay.xlsx not exists!")
    workbook_controller = Workbook()
    workbook_controller.save("Manager'sWorkingDay.xlsx")
    print(f"File with name Manager'sWorkingDay.xlsx successfully created!")

print("File Manager'sWorkingDay.xlsx found!")
workbook_controller = load_workbook("Manager'sWorkingDay.xlsx")

now = datetime.now().strftime("%Y_%m_%d")
try:
    worksheet_controller = workbook_controller[now]
    print(f"Sheet with name {now} found!")
except KeyError:
    print(f"Can't find sheet with name: {now}")
    workbook_controller.create_sheet(now)
    print(f"Sheet with name {now} successfully created!")

workbook_controller.save("Manager'sWorkingDay.xlsx")
workbook_banned_processes = load_workbook("BlockList.xlsx")
worksheet_banned_processes = workbook_banned_processes["BlockList"]

banned_processes = [row[0].value for row in worksheet_banned_processes]
print("BANNED PROCESSES ARE:", *banned_processes)

while True:
    current_banned = set()
    for process in process_iter():
        print("Checking", process.name())

        try:
            if process.name() in banned_processes and process.name() not in current_banned:
                process.kill()
                print(process.name(), "in banned processes!")
                current_banned.add(process.name())

                worksheet_controller.append([datetime.now().strftime("%H:%M:%S"), process.name()])
                workbook_controller.save("Manager'sWorkingDay.xlsx")
        except Exception as e:
            print(e)

    print("Processes had been banned:", *current_banned)
    sleep(10)
