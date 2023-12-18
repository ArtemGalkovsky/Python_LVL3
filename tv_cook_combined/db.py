from sqlite3 import connect, Connection, Cursor
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from os import listdir, getcwd


class DB:
    def __init__(self, table_name: str):
        if table_name not in listdir(getcwd()):
            with open(table_name, "w+", encoding="UTF-8") as fl:
                fl.write("")

        self.connection: Connection = connect(table_name)
        self.cursor: Cursor = self.connection.cursor()
        self.table_path: str = table_name

        self.create_main_db()

    def __del__(self):
        self.connection.close()

    def create_main_db(self):
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS Cooking (
        id TEXT PRIMARY KEY,
        name TEXT,
        category TEXT,
        description TEXT
        );
        """)
        self.connection.commit()

    def delete_table_by_name(self, table_name: str) -> None:
        self.cursor.execute(f"""DROP TABLE '{table_name}';""")
        self.connection.commit()

    def check_if_id_in_db(self, dish_id: str) -> bool:
        self.cursor.execute("SELECT id FROM Cooking WHERE id = ?", (dish_id,))
        cards_with_dish_id = self.cursor.fetchall()

        if cards_with_dish_id:
            return True

        return False

    def add_recipe(self, dish_id: str, ingredients: dict[int: str, ...]) -> None:
        self.cursor.execute(f"""CREATE TABLE IF NOT EXISTS '{dish_id}-recipe' (
            ingredient INTEGER,
            count TEXT
        );""")
        self.connection.commit()

        for ingredient, count in ingredients.items():
            self.cursor.execute(f"INSERT INTO '{dish_id}-recipe' VALUES (?, ?)", (ingredient, count))
        self.connection.commit()

    def add_steps(self, dish_id: str, steps: dict[str: str]):
        self.cursor.execute(f"""CREATE TABLE IF NOT EXISTS '{dish_id}-steps' (
            step_index TEXT,
            step TEXT
        );""")
        self.connection.commit()

        for index, step_description in steps.items():
            self.cursor.execute(f"INSERT INTO '{dish_id}-steps' VALUES (?, ?)", (index, step_description))
        self.connection.commit()

    def add_dish(self, dish: dict[str: str, str: str, str: str, str: str, str: dict[str: str], str: dict[int, str]]) \
            -> None:
        description = dish["description"]
        category = dish["category"]
        name = dish["name"]
        dish_id = dish["id"]
        ingredients = dish["recipe"]
        steps = dish["steps"]

        self.cursor.execute("INSERT INTO Cooking VALUES (?, ?, ?, ?)", (dish_id, name, category, description))
        self.connection.commit()

        self.add_recipe(dish_id, ingredients)
        self.add_steps(dish_id, steps)

    def add_all_dishes(self, all_dishes:
    list[list[dict[str: str, str: str, str: str, str: str, str: dict[str: str], str: dict[int, str]]]]) \
            -> None:
        for page in all_dishes:
            for dish in page:
                dish_id = dish["id"]

                if self.check_if_id_in_db(dish_id):
                    continue

                self.add_dish(dish)

    def delete_all_tables(self):
        with open(self.table_path, "w+", encoding="UTF-8") as fl:
            fl.write("")

    def get_all_dishes_ids(self) -> list[str]:
        self.cursor.execute("SELECT id FROM Cooking")
        return [dish_id[0] for dish_id in self.cursor.fetchall()]

    def add_main_table_to_xlsx(self, dishes_sheet: Worksheet) -> None:
        self.cursor.execute("SELECT * FROM Cooking")
        data = self.cursor.fetchall()

        for dish_id, name, category, description in data:
            dishes_sheet.append((dish_id, name, category, description))

    def add_recipes_to_xlsx(self, recipes_sheet: Worksheet, dishes_ids: list[str]) -> None:
        for dish_id in dishes_ids:
            self.cursor.execute(f"SELECT * FROM '{dish_id}-recipe'")
            recipe = self.cursor.fetchall()

            for ingredient, amount in recipe:
                recipes_sheet.append((dish_id, ingredient, amount))

    def add_steps_to_xlsx(self, steps_sheet: Worksheet, dishes_ids: list[str]) -> None:
        for dish_id in dishes_ids:
            self.cursor.execute(f"SELECT * FROM '{dish_id}-steps'")
            recipe = self.cursor.fetchall()

            for step_index, step_description in recipe:
                steps_sheet.append((dish_id, step_index, step_description))

    def db2xlsx(self) -> None:
        workbook: Workbook = Workbook()
        dishes_sheet: Worksheet = workbook.active
        workbook.create_sheet("Шаги")
        workbook.create_sheet("Рецепты")

        dishes_sheet.title = "Блюда"
        steps_sheet = workbook["Шаги"]
        recipes_sheet = workbook["Рецепты"]

        dishes_sheet.append(("ID", "Название", "Категория", "Описание"))
        steps_sheet.append(("ID", "Шаг", "Шаги"))
        recipes_sheet.append(("ID", "Ингредиенты", "Количество"))

        self.add_main_table_to_xlsx(dishes_sheet)
        self.add_recipes_to_xlsx(recipes_sheet, self.get_all_dishes_ids())
        self.add_steps_to_xlsx(steps_sheet, self.get_all_dishes_ids())

        workbook.save(f"cooking.xlsx")

    def get_dishes_by_category(self, category: str, categories: dict[str]) -> list[tuple[str, str, str, str]]:
        if category in tuple(categories.keys()):
            category = categories[category]
            self.cursor.execute(f"SELECT name, description FROM Cooking WHERE category LIKE ?", ("%" + category + "%",))
            return self.cursor.fetchall()

        else:
            print("[ERROR] Cant get dishes by category: incorrect category name!")

    def get_recipes_by_dishes_name(self, dish_name: str) -> list[list[tuple[str, str]]]:
        self.cursor.execute("SELECT * FROM Cooking WHERE name = ?", (dish_name,))
        dishes_ids = [dish[0] for dish in self.cursor.fetchall()]

        recipes = []

        for dish_id in dishes_ids:
            self.cursor.execute(f"SELECT * FROM '{dish_id}-recipe'")
            recipes.append(self.cursor.fetchall())

        return recipes

    def get_dishes_by_ingredients(self, ingredients: list[str]) -> list[dict[str: str, str: str], ...]:
        self.cursor.execute("SELECT id FROM Cooking")
        dishes_ids = [dish[0] for dish in self.cursor.fetchall()]

        recipes = {}
        for dish_id in dishes_ids:
            self.cursor.execute(f"SELECT * FROM '{dish_id}-recipe'")
            recipes[dish_id] = self.cursor.fetchall()

        dishes = []
        for dish_id, recipe in recipes.items():
            recipe_ingredients_names = [ingredient[0].strip() for ingredient in recipe]

            if all(ingredient_name in recipe_ingredients_names for ingredient_name in ingredients):
                self.cursor.execute(f"SELECT name, description FROM Cooking WHERE id = ?", (dish_id,))
                name_and_description = self.cursor.fetchone()
                dishes.append({"name": name_and_description[0], "description": name_and_description[1]})

        return dishes


if __name__ == "__main__":
    db = DB("cooking.db")
    db.get_dishes_by_ingredients(["Свиная корейка"])
